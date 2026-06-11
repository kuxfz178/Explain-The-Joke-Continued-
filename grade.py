#!/usr/bin/env python3
"""
grade.py reads one or more model run CSVs and writes the Excel grading report.

The output (grading_report.xlsx) has:
  Leaderboard      one row per model, percentages formatted as percentages
  By_Humor_Style   model by humour-style matrix
  By_Context       model by contextual-knowledge matrix
  Per_Item         every (model, item) verdict, for digging into specific cases

Usage:
    python3 grade.py answer_key.csv run1.csv run2.csv ...
"""
import csv
import os
import re
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


LETTERS = ("A", "B", "C", "D")
ANSWER_COL_CANDIDATES = (
    "answer", "model_answer", "response", "choice",
    "prediction", "output", "letter", "pred",
)


# ---------- letter parsing (same logic as stats.py) ----------
def parse_letter(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if s.upper() in LETTERS:
        return s.upper()
    m = re.match(r"^[\s*(\[]*([ABCD])[\s*).:\]]", s.upper() + " ")
    if m:
        return m.group(1)
    hits = re.findall(r"\b([ABCD])\b", s.upper())
    if len(set(hits)) == 1:
        return hits[0]
    return None


def find_answer_col(fieldnames):
    for c in fieldnames:
        if c and c.strip().lower() in ANSWER_COL_CANDIDATES:
            return c
    sys.exit(f"No answer column found in {fieldnames}")


def load_key(path):
    key = {}
    with open(path, encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            key[r["File"].strip()] = r
    return key


def score_model(path, key):
    with open(path, encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        fields = reader.fieldnames or []
    if "File" not in fields:
        sys.exit(f"{path}: no 'File' column")
    acol = find_answer_col(fields)
    model = (rows[0].get("model") or "").strip() if rows else ""
    if not model:
        model = os.path.splitext(os.path.basename(path))[0]

    results = []
    for r in rows:
        fid = r["File"].strip()
        k = key.get(fid)
        if k is None:
            continue
        letter = parse_letter(r.get(acol))
        correct = k["correct_letter"].strip().upper()
        literal = k["literal_letter"].strip().upper()
        if letter is None:
            status = "invalid"
        elif letter == correct:
            status = "correct"
        elif literal != "UNKNOWN" and letter == literal:
            status = "literal"
        elif literal == "UNKNOWN":
            status = "wrong_literal_unknown"
        else:
            status = "other_distractor"
        results.append({
            "File": fid,
            "model": model,
            "model_raw_answer": (r.get(acol) or "").strip(),
            "parsed_letter": letter or "",
            "correct_letter": correct,
            "literal_letter": literal,
            "status": status,
            "humor_style": k.get("humor_style", ""),
            "humor_style_secondary": k.get("humor_style_secondary", ""),
            "contextual_knowledge": k.get("contextual_knowledge", ""),
            "contextual_knowledge_secondary":
                k.get("contextual_knowledge_secondary", ""),
        })
    return model, results


def aggregate(results):
    n = len(results)
    correct = sum(r["status"] == "correct" for r in results)
    literal = sum(r["status"] == "literal" for r in results)
    other = sum(r["status"] == "other_distractor" for r in results)
    invalid = sum(r["status"] == "invalid" for r in results)
    wlu = sum(r["status"] == "wrong_literal_unknown" for r in results)
    lit_known = sum(
        1 for r in results
        if r["literal_letter"] != "UNKNOWN" and r["status"] != "invalid"
    )
    errors_known = literal + other
    return {
        "n": n,
        "correct": correct,
        "literal": literal,
        "other_distractor": other,
        "wrong_literal_unknown": wlu,
        "invalid": invalid,
        "accuracy": correct / n if n else 0,
        "accuracy_excl_invalid": correct / (n - invalid) if n - invalid else 0,
        "format_compliance": (n - invalid) / n if n else 0,
        "literal_trap_rate": literal / lit_known if lit_known else 0,
        "literal_share_of_errors": (
            literal / errors_known if errors_known else 0
        ),
    }


def by_category(results, field):
    """Breakdown per category with the full status counts and the derived rates."""
    agg = defaultdict(lambda: {"n": 0, "correct": 0, "literal": 0,
                                "other": 0, "wlu": 0, "invalid": 0,
                                "lit_known": 0})
    for r in results:
        # Multi-label safe. We split on ; or , but NOT on /, because a few of
        # our category names contain a slash ("Sarcasm/Irony", "History/Mythology").
        cats = [c.strip() for c in re.split(r"[;,]", r[field]) if c.strip()]
        for cat in cats or ["(blank)"]:
            a = agg[cat]
            a["n"] += 1
            s = r["status"]
            a["correct"] += s == "correct"
            a["literal"] += s == "literal"
            a["other"]   += s == "other_distractor"
            a["wlu"]     += s == "wrong_literal_unknown"
            a["invalid"] += s == "invalid"
            if r["literal_letter"] != "UNKNOWN" and s != "invalid":
                a["lit_known"] += 1
    out = {}
    for cat, a in agg.items():
        errors_known = a["literal"] + a["other"]
        out[cat] = {
            "n": a["n"],
            "correct": a["correct"],
            "literal": a["literal"],
            "other_distractor": a["other"],
            "wrong_literal_unknown": a["wlu"],
            "invalid": a["invalid"],
            "accuracy": a["correct"] / a["n"] if a["n"] else 0,
            "literal_trap_rate": (
                a["literal"] / a["lit_known"] if a["lit_known"] else 0
            ),
            "literal_share_of_errors": (
                a["literal"] / errors_known if errors_known else 0
            ),
        }
    return out


# ---------- Excel writers ----------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FMT = "0.0%"
INT_FMT = "0"


def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER


def auto_width(ws, headers, data_rows):
    for i, h in enumerate(headers, 1):
        width = max(len(str(h)), 8)
        for row in data_rows:
            v = row[i - 1] if i - 1 < len(row) else ""
            width = max(width, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 50)


def write_leaderboard(wb, models_data):
    """Writes the Leaderboard sheet. models_data is a list of per-model dicts."""
    ws = wb.create_sheet("Leaderboard")
    headers = [
        "Model", "Items", "Correct", "Accuracy", "Accuracy excl. invalid",
        "Literal-trap rate", "Literal share of errors",
        "Other distractor", "Invalid",
    ]
    write_header(ws, 1, headers)

    # sort by accuracy descending
    rows_sorted = sorted(
        models_data, key=lambda m: -m["overall"]["accuracy"]
    )
    for i, m in enumerate(rows_sorted, start=2):
        model = m["model"]; o = m["overall"]
        ws.cell(row=i, column=1, value=model).border = BORDER
        ws.cell(row=i, column=2, value=o["n"]).number_format = INT_FMT
        ws.cell(row=i, column=3, value=o["correct"]).number_format = INT_FMT
        c = ws.cell(row=i, column=4, value=o["accuracy"])
        c.number_format = PCT_FMT
        c = ws.cell(row=i, column=5, value=o["accuracy_excl_invalid"])
        c.number_format = PCT_FMT
        c = ws.cell(row=i, column=6, value=o["literal_trap_rate"])
        c.number_format = PCT_FMT
        c = ws.cell(row=i, column=7, value=o["literal_share_of_errors"])
        c.number_format = PCT_FMT
        ws.cell(row=i, column=8,
                value=o["other_distractor"]).number_format = INT_FMT
        ws.cell(row=i, column=9, value=o["invalid"]).number_format = INT_FMT
        # zebra-stripe + borders
        for col in range(1, len(headers) + 1):
            cc = ws.cell(row=i, column=col)
            cc.border = BORDER
            if i % 2 == 0:
                cc.fill = ALT_FILL
    ws.freeze_panes = "B2"
    auto_width(ws, headers, [[m["model"]] + [""] * (len(headers) - 1)
                              for m in rows_sorted])


def write_matrix(wb, sheet_name, models_data, field):
    """Writes a model-by-category matrix sheet.

    sheet_name   the tab name, for example By_Humor_Style or By_Humor_Secondary.
    field        which per-model aggregate to pull (humor_style,
                 humor_style_secondary, contextual_knowledge, or
                 contextual_knowledge_secondary).
    models_data  list of per-model dicts.
    """
    ws = wb.create_sheet(sheet_name)

    # Collect every category any model used and sort them.
    all_cats = set()
    for m in models_data:
        all_cats.update(m[field].keys())
    cats = sorted(all_cats)

    # n is the same for each category across models, so grab it from the first.
    n_per_cat = {}
    if models_data:
        agg0 = models_data[0][field]
        for cat in cats:
            n_per_cat[cat] = agg0.get(cat, {}).get("n", 0)

    # Two side by side blocks: accuracy on the left, literal-trap rate on the right.
    # Row 1 is the block heading.
    ws.cell(row=1, column=1, value="Accuracy by category").font = Font(
        bold=True, size=12, color="4472C4"
    )
    ws.cell(row=1, column=len(cats) + 3,
            value="Literal-trap rate by category").font = Font(
        bold=True, size=12, color="C00000"
    )

    # Row 2 holds the category name, row 3 the item count.
    col = 1
    ws.cell(row=2, column=col, value="Model").font = HEADER_FONT
    ws.cell(row=2, column=col).fill = HEADER_FILL
    ws.cell(row=2, column=col).border = BORDER
    ws.cell(row=3, column=col, value="").border = BORDER
    col += 1
    for cat in cats:
        c = ws.cell(row=2, column=col, value=cat)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
        sub = ws.cell(row=3, column=col, value=f"n={n_per_cat.get(cat, 0)}")
        sub.font = Font(italic=True, size=9, color="595959")
        sub.alignment = Alignment(horizontal="center")
        sub.border = BORDER
        col += 1
    # Empty column between the two blocks.
    col += 1
    # Second block headers.
    for cat in cats:
        c = ws.cell(row=2, column=col, value=cat)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
        sub = ws.cell(row=3, column=col, value=f"n={n_per_cat.get(cat, 0)}")
        sub.font = Font(italic=True, size=9, color="595959")
        sub.alignment = Alignment(horizontal="center")
        sub.border = BORDER
        col += 1

    # Data starts on row 4 since the headers take rows 2 and 3.
    rows_sorted = sorted(models_data, key=lambda m: -m["overall"]["accuracy"])
    for r_idx, m in enumerate(rows_sorted, start=4):
        model = m["model"]
        agg = m[field]
        col = 1
        ws.cell(row=r_idx, column=col, value=model).border = BORDER
        if r_idx % 2 == 1:
            ws.cell(row=r_idx, column=col).fill = ALT_FILL
        col += 1
        for cat in cats:
            v = agg.get(cat, {}).get("accuracy", "")
            c = ws.cell(row=r_idx, column=col,
                        value=v if v != "" else None)
            if v != "":
                c.number_format = PCT_FMT
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
            if r_idx % 2 == 1:
                c.fill = ALT_FILL
            col += 1
        # Skip the empty column between blocks.
        col += 1
        for cat in cats:
            v = agg.get(cat, {}).get("literal_trap_rate", "")
            c = ws.cell(row=r_idx, column=col,
                        value=v if v != "" else None)
            if v != "":
                c.number_format = PCT_FMT
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
            if r_idx % 2 == 1:
                c.fill = ALT_FILL
            col += 1

    ws.freeze_panes = "B4"
    ws.row_dimensions[2].height = 28   # room for the bold category name
    ws.column_dimensions["A"].width = 36
    for i in range(2, col):
        ws.column_dimensions[get_column_letter(i)].width = 14
    # The empty column between blocks should be narrow.
    ws.column_dimensions[get_column_letter(len(cats) + 2)].width = 2


def write_details(wb, sheet_name, models_data, field):
    """Long table with one row per (model, category).

    Columns are Model, Category, n, correct, literal, other_distractor,
    invalid, accuracy, literal_trap_rate, and literal_share_of_errors.

    field picks which aggregate to write (humor_style, humor_style_secondary,
    contextual_knowledge, or contextual_knowledge_secondary).
    """
    ws = wb.create_sheet(sheet_name)
    headers = [
        "Model", "Category", "n", "Correct", "Literal", "Other distractor",
        "Invalid", "Accuracy", "Literal-trap rate", "Literal share of errors",
    ]
    write_header(ws, 1, headers)

    row = 2
    rows_sorted = sorted(models_data, key=lambda m: -m["overall"]["accuracy"])
    for m in rows_sorted:
        model = m["model"]
        agg = m[field]
        # Sort categories alphabetically inside each model block.
        for cat in sorted(agg.keys()):
            a = agg[cat]
            ws.cell(row=row, column=1, value=model)
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=a["n"]).number_format = INT_FMT
            ws.cell(row=row, column=4,
                    value=a["correct"]).number_format = INT_FMT
            ws.cell(row=row, column=5,
                    value=a["literal"]).number_format = INT_FMT
            ws.cell(row=row, column=6,
                    value=a["other_distractor"]).number_format = INT_FMT
            ws.cell(row=row, column=7,
                    value=a["invalid"]).number_format = INT_FMT
            c = ws.cell(row=row, column=8, value=a["accuracy"])
            c.number_format = PCT_FMT
            c = ws.cell(row=row, column=9, value=a["literal_trap_rate"])
            c.number_format = PCT_FMT
            c = ws.cell(row=row, column=10,
                        value=a["literal_share_of_errors"])
            c.number_format = PCT_FMT
            # Borders and zebra striping.
            for col in range(1, 11):
                cc = ws.cell(row=row, column=col)
                cc.border = BORDER
                if row % 2 == 0:
                    cc.fill = ALT_FILL
            row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [36, 24, 6, 8, 8, 12, 8, 11, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_difficulty(wb, sheet_name, models_data, field):
    """One row per category showing how hard it was across all models.

    Columns are Category, n, average accuracy, best model, best accuracy,
    worst model, worst accuracy, the spread between them, and the average
    literal-trap rate for that category.
    """
    ws = wb.create_sheet(sheet_name)
    headers = [
        "Category", "n",
        "Avg accuracy (across models)", "Best model", "Best accuracy",
        "Worst model", "Worst accuracy", "Range (best - worst)",
        "Avg literal-trap rate",
    ]
    write_header(ws, 1, headers)

    # Group the per-category numbers across models.
    by_cat = defaultdict(list)   # cat -> [(model, accuracy, literal_rate, n)]
    n_by_cat = {}
    for m in models_data:
        model = m["model"]
        agg = m[field]
        for cat, a in agg.items():
            by_cat[cat].append(
                (model, a["accuracy"], a["literal_trap_rate"], a["n"])
            )
            n_by_cat[cat] = a["n"]

    # Sort categories so the hardest ones are at the top.
    cats = list(by_cat.keys())
    def avg_acc(cat):
        accs = [acc for (_, acc, _, _) in by_cat[cat]]
        return sum(accs) / len(accs) if accs else 0
    cats.sort(key=avg_acc)

    row = 2
    for cat in cats:
        entries = by_cat[cat]
        accs = [acc for (_, acc, _, _) in entries]
        lits = [lit for (_, _, lit, _) in entries]
        avg_a = sum(accs) / len(accs)
        avg_l = sum(lits) / len(lits)
        best  = max(entries, key=lambda e: e[1])
        worst = min(entries, key=lambda e: e[1])

        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2,
                value=n_by_cat[cat]).number_format = INT_FMT
        c = ws.cell(row=row, column=3, value=avg_a); c.number_format = PCT_FMT
        ws.cell(row=row, column=4, value=best[0])
        c = ws.cell(row=row, column=5, value=best[1]); c.number_format = PCT_FMT
        ws.cell(row=row, column=6, value=worst[0])
        c = ws.cell(row=row, column=7, value=worst[1]); c.number_format = PCT_FMT
        c = ws.cell(row=row, column=8, value=best[1] - worst[1])
        c.number_format = PCT_FMT
        c = ws.cell(row=row, column=9, value=avg_l); c.number_format = PCT_FMT
        for col in range(1, 10):
            cc = ws.cell(row=row, column=col)
            cc.border = BORDER
            if row % 2 == 0:
                cc.fill = ALT_FILL
        row += 1

    ws.freeze_panes = "A2"
    widths = [24, 6, 18, 32, 12, 32, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_per_item(wb, all_results):
    """Every (model, item) verdict in one filterable sheet."""
    ws = wb.create_sheet("Per_Item")
    headers = [
        "Model", "File", "Status", "Parsed letter",
        "Correct letter", "Literal letter",
        "Humor style", "Contextual knowledge",
        "Raw model answer",
    ]
    write_header(ws, 1, headers)
    for i, r in enumerate(all_results, start=2):
        ws.cell(row=i, column=1, value=r["model"]).border = BORDER
        ws.cell(row=i, column=2, value=r["File"]).border = BORDER
        c = ws.cell(row=i, column=3, value=r["status"])
        c.border = BORDER
        # Colour the row by status so the wrong answers are easy to spot.
        colors = {
            "correct": "C6EFCE", "literal": "FFC7CE",
            "other_distractor": "FFEB9C",
            "invalid": "D9D9D9", "wrong_literal_unknown": "D9D9D9",
        }
        if r["status"] in colors:
            c.fill = PatternFill("solid", fgColor=colors[r["status"]])
        ws.cell(row=i, column=4, value=r["parsed_letter"]).border = BORDER
        ws.cell(row=i, column=5, value=r["correct_letter"]).border = BORDER
        ws.cell(row=i, column=6, value=r["literal_letter"]).border = BORDER
        ws.cell(row=i, column=7, value=r["humor_style"]).border = BORDER
        ws.cell(row=i, column=8,
                value=r["contextual_knowledge"]).border = BORDER
        ws.cell(row=i, column=9,
                value=r["model_raw_answer"][:200]).border = BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [32, 12, 22, 8, 8, 8, 18, 22, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- main ----------
def main():
    if len(sys.argv) < 3:
        sys.exit(
            "usage: python3 grade.py answer_key.csv run1.csv [run2.csv ...]"
        )
    key_path = sys.argv[1]
    run_paths = sys.argv[2:]

    key = load_key(key_path)
    print(f"answer key: {len(key)} items from {key_path}")

    models_data = []
    all_results = []
    for path in run_paths:
        model, results = score_model(path, key)
        if not results:
            print(f"  {path}: no rows graded, skipping")
            continue
        overall = aggregate(results)
        models_data.append({
            "model": model,
            "overall": overall,
            "humor_style":             by_category(results, "humor_style"),
            "humor_style_secondary":   by_category(results,
                                                  "humor_style_secondary"),
            "contextual_knowledge":    by_category(results,
                                                   "contextual_knowledge"),
            "contextual_knowledge_secondary": by_category(
                results, "contextual_knowledge_secondary"),
        })
        all_results.extend(results)
        print(f"  {model:42s}  n={overall['n']:4d}  "
              f"acc={overall['accuracy']:5.1%}  "
              f"literal={overall['literal_trap_rate']:5.1%}  "
              f"invalid={overall['invalid']}")

    # Build the workbook.
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_leaderboard(wb, models_data)

    # Primary humour and context breakdowns.
    write_matrix(wb, "By_Humor_Style", models_data, "humor_style")
    write_matrix(wb, "By_Context",     models_data, "contextual_knowledge")
    write_details(wb,    "Humor_Details",    models_data, "humor_style")
    write_details(wb,    "Context_Details",  models_data, "contextual_knowledge")
    write_difficulty(wb, "Humor_Difficulty", models_data, "humor_style")
    write_difficulty(wb, "Context_Difficulty", models_data,
                     "contextual_knowledge")

    # Secondary humour and context breakdowns.
    write_matrix(wb, "By_Humor_Secondary",   models_data,
                 "humor_style_secondary")
    write_matrix(wb, "By_Context_Secondary", models_data,
                 "contextual_knowledge_secondary")
    write_details(wb,    "Humor_Sec_Details",    models_data,
                  "humor_style_secondary")
    write_details(wb,    "Context_Sec_Details",  models_data,
                  "contextual_knowledge_secondary")
    write_difficulty(wb, "Humor_Sec_Difficulty", models_data,
                     "humor_style_secondary")
    write_difficulty(wb, "Context_Sec_Difficulty", models_data,
                     "contextual_knowledge_secondary")

    write_per_item(wb, all_results)

    out_path = "grading_report.xlsx"
    wb.save(out_path)
    print(f"\nwrote {out_path}")
    print("  open it in Excel or Sheets. The percentage cells are formatted "
          "as percentages so you should not see anything like 146:09:00.")


if __name__ == "__main__":
    main()
