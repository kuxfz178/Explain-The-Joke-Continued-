#!/usr/bin/env python3
"""
Layers the secondary humour and context labels from the updated annotation
spreadsheet onto answer_key.csv without touching the existing correct or
literal letters. We needed this once the secondary annotations were added in
the second pass, so we could keep the existing model runs valid.
"""
import csv
import openpyxl


NEW_SRC = "picture_annotations_v2.xlsx"
KEY_CSV = "answer_key.csv"

# Pull the secondary columns from the new source (col 10 humour, col 12 context).
wb = openpyxl.load_workbook(NEW_SRC, data_only=True)
ws = wb["Annotations"]
new_sec = {}
for r in range(2, ws.max_row + 1):
    fid = ws.cell(r, 1).value
    if not fid:
        continue
    new_sec[str(fid).strip()] = {
        "humor_style_secondary": (ws.cell(r, 10).value or "").strip(),
        "contextual_knowledge_secondary": (ws.cell(r, 12).value or "").strip(),
    }
print(f"loaded {len(new_sec)} secondary annotations from {NEW_SRC}")

# Read the existing answer_key so we can layer the new labels on top.
with open(KEY_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    rows = list(reader)
    fields = list(reader.fieldnames)

# If the secondary columns are not in the answer_key yet, add them right after
# the matching primary column.
for new_col in ("humor_style_secondary", "contextual_knowledge_secondary"):
    if new_col not in fields:
        if new_col == "humor_style_secondary":
            i = fields.index("humor_style") + 1
        else:
            i = fields.index("contextual_knowledge") + 1
        fields.insert(i, new_col)

overlaid = 0
not_in_new = 0
for row in rows:
    fid = row["File"]
    if fid in new_sec:
        row["humor_style_secondary"] = new_sec[fid]["humor_style_secondary"]
        row["contextual_knowledge_secondary"] = (
            new_sec[fid]["contextual_knowledge_secondary"])
        overlaid += 1
    else:
        # If a row is not in the new source, leave the secondary blank.
        row.setdefault("humor_style_secondary", "")
        row.setdefault("contextual_knowledge_secondary", "")
        not_in_new += 1

with open(KEY_CSV, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
    w.writeheader()
    w.writerows(rows)

print(f"overlaid {overlaid} rows with the new secondary labels")
print(f"{not_in_new} rows in the old source had no match in the new source (left blank)")
print(f"wrote the secondary columns to {KEY_CSV}")
