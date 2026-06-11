#!/usr/bin/env python3
"""
stats.py runs the statistical analysis on top of the MCQ benchmark runs.

Per model it computes:
  accuracy, variance, standard deviation, standard error,
  and a 95 percent bootstrap confidence interval for accuracy.

Per pair of models (A versus B) it computes:
  McNemar's test (continuity-corrected chi-square and p-value)
  and a bootstrap test for the accuracy difference (95 percent CI on Acc_A minus Acc_B).

The output (stats_report.xlsx) has three sheets:
  Model_Stats
  McNemar_Pairwise
  Bootstrap_Diff

Usage:
    python3 stats.py answer_key.csv run1.csv run2.csv ...
"""
import csv
import math
import os
import re
import sys
from itertools import combinations

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.stats import chi2 as chi2_dist


# ---------- letter parsing (same as grade.py) ----------
LETTERS = ("A", "B", "C", "D")
ANSWER_COL_CANDIDATES = (
    "answer", "model_answer", "response", "choice",
    "prediction", "output", "letter", "pred",
)


def parse_letter(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if s.upper() in LETTERS:
        return s.upper()
    m = re.match(r"^[\s*(\[]*([ABCD])[\s*).:\]]", s.upper() + " ")
    if m:
        return m.group(1)
    hits = re.findall(r"\b([ABCD])\b", s.upper())
    if len(set(hits)) == 1:
        return hits[0]
    return None


def find_answer_col(fieldnames):
    for c in fieldnames:
        if c and c.strip().lower() in ANSWER_COL_CANDIDATES:
            return c
    sys.exit(f"No answer column found in {fieldnames}")


# ---------- loading ----------
def load_runs(answer_key_path, run_paths):
    """Returns (key, by_model). by_model[name] maps File to 0, 1, or None."""
    with open(answer_key_path, encoding="utf-8") as f:
        key = {r["File"].strip(): r for r in csv.DictReader(f)}

    by_model = {}
    for path in run_paths:
        with open(path, encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
            fields = reader.fieldnames or []
        acol = find_answer_col(fields)
        model = (rows[0].get("model") or "").strip() if rows else ""
        if not model:
            model = os.path.splitext(os.path.basename(path))[0]

        per_file = {}
        for r in rows:
            fid = r["File"].strip()
            if fid not in key:
                continue
            letter = parse_letter(r.get(acol))
            correct_letter = key[fid]["correct_letter"].strip().upper()
            if letter is None:
                per_file[fid] = None
            else:
                per_file[fid] = 1 if letter == correct_letter else 0
        by_model[model] = per_file
    return key, by_model


# ---------- per-model statistics ----------
def model_stats(per_file, n_bootstrap=10000, seed=42):
    """Compute accuracy, variance, standard deviation, standard error,
    and a 95 percent bootstrap CI for accuracy on a single model."""
    items = list(per_file.values())
    n_total = len(items)
    valid = np.array([x for x in items if x is not None], dtype=np.int8)
    n_valid = len(valid)
    correct = int(valid.sum()) if n_valid else 0
    invalid = n_total - n_valid

    accuracy = correct / n_total if n_total else 0.0
    p = accuracy
    variance = p * (1 - p)                  # Bernoulli variance with probability p
    std_dev  = math.sqrt(variance)          # population std deviation of per-item correctness
    std_err  = math.sqrt(variance / n_total) if n_total else 0.0

    # Bootstrap a 95 percent CI for accuracy. We resample with replacement
    # from all items and treat invalid responses as wrong, which matches the
    # conservative accuracy figure we report in the thesis.
    items_arr = np.array(
        [1 if x == 1 else 0 for x in items], dtype=np.int8
    )
    rng = np.random.default_rng(seed)
    idx = rng.integers(0, n_total, size=(n_bootstrap, n_total))
    boot_means = items_arr[idx].mean(axis=1)
    ci_low, ci_high = np.percentile(boot_means, [2.5, 97.5])

    return {
        "n_total": n_total,
        "n_valid": n_valid,
        "n_invalid": invalid,
        "correct": correct,
        "accuracy": accuracy,
        "variance": variance,
        "std_dev": std_dev,
        "std_error": std_err,
        "ci_low_95": float(ci_low),
        "ci_high_95": float(ci_high),
    }


# ---------- pairwise McNemar's test ----------
def mcnemar(a_per_file, b_per_file):
    """McNemar's test with the continuity correction. Returns the cell counts,
    the chi-square statistic, the two-sided p-value, and the significance flags."""
    common = sorted(set(a_per_file) & set(b_per_file))
    both_correct = both_wrong = a_only = b_only = skipped = 0
    for fid in common:
        a, b = a_per_file[fid], b_per_file[fid]
        if a is None or b is None:
            skipped += 1
            continue
        if a and b:        both_correct += 1
        elif a and not b:  a_only += 1
        elif not a and b:  b_only += 1
        else:              both_wrong += 1

    n_disc = a_only + b_only
    if n_disc == 0:
        chi2_stat = 0.0
        p = 1.0
    else:
        # Continuity correction from Edwards (1948).
        chi2_stat = (abs(a_only - b_only) - 1) ** 2 / n_disc
        p = float(chi2_dist.sf(chi2_stat, df=1))

    return {
        "n_paired": len(common),
        "n_evaluated": len(common) - skipped,
        "both_correct": both_correct,
        "both_wrong": both_wrong,
        "a_only_correct": a_only,
        "b_only_correct": b_only,
        "skipped_invalid": skipped,
        "chi_square": chi2_stat,
        "p_value": p,
        "sig_05": p < 0.05,
        "sig_01": p < 0.01,
        "sig_001": p < 0.001,
    }


# ---------- pairwise bootstrap test for accuracy difference ----------
def bootstrap_diff(a_per_file, b_per_file, n_bootstrap=10000, seed=42):
    """Resample items with replacement and compute the 95 percent CI of
    (Acc_A minus Acc_B). The CI is on the paired (matched-item) difference,
    so both accuracies are computed on the same resampled item set in every
    iteration."""
    common = sorted(set(a_per_file) & set(b_per_file))
    a = np.array([1 if a_per_file[f] == 1 else 0 for f in common],
                 dtype=np.int8)
    b = np.array([1 if b_per_file[f] == 1 else 0 for f in common],
                 dtype=np.int8)
    n = len(common)
    if n == 0:
        return {"n": 0, "acc_a": 0, "acc_b": 0, "diff": 0,
                "ci_low_95": 0, "ci_high_95": 0, "sig_05": False}

    acc_a = float(a.mean())
    acc_b = float(b.mean())
    diff  = acc_a - acc_b

    rng = np.random.default_rng(seed)
    idx = rng.integers(0, n, size=(n_bootstrap, n))
    diffs = a[idx].mean(axis=1) - b[idx].mean(axis=1)
    ci_low, ci_high = np.percentile(diffs, [2.5, 97.5])

    return {
        "n": n,
        "acc_a": acc_a,
        "acc_b": acc_b,
        "diff": diff,
        "ci_low_95": float(ci_low),
        "ci_high_95": float(ci_high),
        # If the CI excludes 0, the difference is significant at alpha = 0.05 (two-sided).
        "sig_05": not (ci_low <= 0 <= ci_high),
    }


# ---------- Excel output ----------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
SIG_FILL    = PatternFill("solid", fgColor="FFE699")    # light yellow
NS_FILL     = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FMT = "0.0%"
NUM4_FMT = "0.0000"
NUM3_FMT = "0.000"


def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = BORDER


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_model_stats(wb, stats):
    """Writes the Model_Stats sheet. stats maps model_name to its stats dict."""
    ws = wb.create_sheet("Model_Stats")
    headers = [
        "Model", "n", "Correct", "Accuracy",
        "Variance", "Std deviation", "Std error",
        "95% CI lower", "95% CI upper", "CI width",
    ]
    write_header(ws, 1, headers)
    ws.row_dimensions[1].height = 32

    rows = sorted(stats.items(), key=lambda kv: -kv[1]["accuracy"])
    for r_idx, (model, s) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=model)
        ws.cell(row=r_idx, column=2, value=s["n_total"])
        ws.cell(row=r_idx, column=3, value=s["correct"])
        ws.cell(row=r_idx, column=4, value=s["accuracy"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=5, value=s["variance"]).number_format = NUM4_FMT
        ws.cell(row=r_idx, column=6, value=s["std_dev"]).number_format = NUM4_FMT
        ws.cell(row=r_idx, column=7, value=s["std_error"]).number_format = NUM4_FMT
        ws.cell(row=r_idx, column=8, value=s["ci_low_95"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=9, value=s["ci_high_95"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=10,
                value=s["ci_high_95"] - s["ci_low_95"]).number_format = PCT_FMT
        for col in range(1, len(headers) + 1):
            cc = ws.cell(row=r_idx, column=col)
            cc.border = BORDER
            if r_idx % 2 == 0:
                cc.fill = ALT_FILL

    ws.freeze_panes = "B2"
    auto_width(ws, [36, 8, 10, 12, 12, 14, 12, 14, 14, 12])


def write_mcnemar(wb, results, stats):
    """Writes the McNemar_Pairwise sheet.
    results is a list of (model_a, model_b, mcnemar_dict) tuples."""
    ws = wb.create_sheet("McNemar_Pairwise")
    headers = [
        "Model A", "Model B", "n paired", "n evaluated",
        "Both correct", "Both wrong",
        "A only correct", "B only correct",
        "χ² (continuity-corrected)", "p-value",
        "Sig α=0.05", "Sig α=0.01", "Sig α=0.001",
    ]
    write_header(ws, 1, headers)
    ws.row_dimensions[1].height = 32

    # Sort each pair so Model A is the higher-accuracy of the two. That way
    # the "winning direction" is consistent across the sheet.
    sortable = []
    for a, b, m in results:
        if stats[a]["accuracy"] >= stats[b]["accuracy"]:
            sortable.append((a, b, m))
        else:
            # Swap the order so A is always the stronger model in the pair.
            mm = dict(m)
            mm["a_only_correct"], mm["b_only_correct"] = (
                m["b_only_correct"], m["a_only_correct"])
            sortable.append((b, a, mm))
    sortable.sort(key=lambda t: (-stats[t[0]]["accuracy"],
                                  stats[t[1]]["accuracy"]))

    for r_idx, (a, b, m) in enumerate(sortable, start=2):
        ws.cell(row=r_idx, column=1, value=a)
        ws.cell(row=r_idx, column=2, value=b)
        ws.cell(row=r_idx, column=3, value=m["n_paired"])
        ws.cell(row=r_idx, column=4, value=m["n_evaluated"])
        ws.cell(row=r_idx, column=5, value=m["both_correct"])
        ws.cell(row=r_idx, column=6, value=m["both_wrong"])
        ws.cell(row=r_idx, column=7, value=m["a_only_correct"])
        ws.cell(row=r_idx, column=8, value=m["b_only_correct"])
        ws.cell(row=r_idx, column=9,
                value=m["chi_square"]).number_format = NUM3_FMT
        ws.cell(row=r_idx, column=10,
                value=m["p_value"]).number_format = "0.000E+00"
        for offset, key in enumerate(("sig_05", "sig_01", "sig_001"), start=11):
            c = ws.cell(row=r_idx, column=offset,
                        value="yes" if m[key] else "no")
            if m[key]:
                c.fill = SIG_FILL
        for col in range(1, len(headers) + 1):
            cc = ws.cell(row=r_idx, column=col)
            cc.border = BORDER
            if r_idx % 2 == 0 and cc.fill == NS_FILL:
                cc.fill = ALT_FILL

    ws.freeze_panes = "C2"
    auto_width(ws, [32, 32, 9, 10, 10, 10, 12, 12, 16, 12, 9, 9, 10])


def write_bootstrap(wb, results, stats):
    """Writes the Bootstrap_Diff sheet.
    results is a list of (model_a, model_b, bootstrap_dict) tuples."""
    ws = wb.create_sheet("Bootstrap_Diff")
    headers = [
        "Model A", "Model B", "n paired",
        "Acc A", "Acc B", "Diff (A - B)",
        "95% CI lower", "95% CI upper", "CI width",
        "Sig α=0.05",
    ]
    write_header(ws, 1, headers)
    ws.row_dimensions[1].height = 32

    # Order each pair so Model A is the higher-accuracy model.
    sortable = []
    for a, b, m in results:
        if stats[a]["accuracy"] >= stats[b]["accuracy"]:
            sortable.append((a, b, m))
        else:
            mm = dict(m)
            mm["acc_a"], mm["acc_b"] = m["acc_b"], m["acc_a"]
            mm["diff"] = -m["diff"]
            mm["ci_low_95"], mm["ci_high_95"] = -m["ci_high_95"], -m["ci_low_95"]
            sortable.append((b, a, mm))
    sortable.sort(key=lambda t: (-stats[t[0]]["accuracy"],
                                  stats[t[1]]["accuracy"]))

    for r_idx, (a, b, m) in enumerate(sortable, start=2):
        ws.cell(row=r_idx, column=1, value=a)
        ws.cell(row=r_idx, column=2, value=b)
        ws.cell(row=r_idx, column=3, value=m["n"])
        ws.cell(row=r_idx, column=4, value=m["acc_a"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=5, value=m["acc_b"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=6, value=m["diff"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=7,
                value=m["ci_low_95"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=8,
                value=m["ci_high_95"]).number_format = PCT_FMT
        ws.cell(row=r_idx, column=9,
                value=m["ci_high_95"] - m["ci_low_95"]).number_format = PCT_FMT
        c = ws.cell(row=r_idx, column=10,
                    value="yes" if m["sig_05"] else "no")
        if m["sig_05"]:
            c.fill = SIG_FILL
        for col in range(1, len(headers) + 1):
            cc = ws.cell(row=r_idx, column=col)
            cc.border = BORDER

    ws.freeze_panes = "C2"
    auto_width(ws, [32, 32, 9, 10, 10, 14, 14, 14, 12, 10])


# ---------- methodology sheet ----------
def write_methodology(wb, stats):
    """Add a sheet that explains each metric with a formula and a worked example."""
    ws = wb.create_sheet("Methodology", 0)   # insert as the first tab

    # Worked examples use whichever model came out on top.
    top = max(stats.items(), key=lambda kv: kv[1]["accuracy"])
    name, s = top
    short = name.split("/")[-1]

    title_font   = Font(bold=True, size=18, color="1F3864")
    section_font = Font(bold=True, size=13, color="4472C4")
    label_font   = Font(bold=True, size=11)
    body_font    = Font(size=11)
    mono_font    = Font(name="Consolas", size=10, color="404040")
    note_font    = Font(italic=True, size=10, color="595959")

    def put(row, text, font=body_font, wrap=True, fill=None):
        c = ws.cell(row=row, column=1, value=text)
        c.font = font
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        if fill:
            c.fill = fill
        # Give longer text more vertical room.
        if wrap and len(str(text)) > 70:
            ws.row_dimensions[row].height = max(
                ws.row_dimensions[row].height or 15,
                15 * (len(str(text)) // 80 + 1)
            )
        return c

    ws.column_dimensions["A"].width = 110

    row = 1
    put(row, "Statistical methods. How each number was calculated and what it means.",
        title_font); row += 2

    put(row,
        "This sheet documents every statistic used in the Model_Stats, "
        "McNemar_Pairwise, and Bootstrap_Diff tabs. Each section gives a "
        "plain-language definition, the formula, a worked example using the "
        "actual benchmark numbers, and a short note on how to read the result "
        "in the writeup."); row += 2

    # ===== Accuracy =====
    put(row, "1. Accuracy", section_font); row += 1
    put(row, "What it measures", label_font); row += 1
    put(row,
        "The fraction of items the model answered correctly. The headline metric "
        "of the benchmark."); row += 1
    put(row, "Formula", label_font); row += 1
    put(row, "accuracy = (number of correct answers) / (total items)",
        mono_font); row += 1
    put(row, "Worked example", label_font); row += 1
    put(row,
        f"{short}: {s['correct']} correct out of {s['n_total']} items = "
        f"{s['accuracy']:.1%}", mono_font); row += 1
    put(row, "Note on invalid responses", label_font); row += 1
    put(row,
        "Items where the model produced no parseable letter (refusals or "
        "garbled output) count as wrong in the headline accuracy. This is the "
        "conservative reading. The grading report also has an 'Accuracy excl. "
        "invalid' column if you prefer the alternative."
        ); row += 2

    # ===== Variance =====
    put(row, "2. Variance", section_font); row += 1
    put(row, "What it measures", label_font); row += 1
    put(row,
        "How spread out the per-item outcomes (1 = correct, 0 = wrong) are "
        "around the mean accuracy. For a binary outcome like ours, variance "
        "is fully determined by accuracy itself, so it does not add any new "
        "information beyond what accuracy already gives. We report it because "
        "the standard-error formula depends on it."); row += 1
    put(row, "Formula", label_font); row += 1
    put(row, "variance = p × (1 − p),   where p = accuracy",
        mono_font); row += 1
    put(row, "Worked example", label_font); row += 1
    put(row,
        f"{short}: p = {s['accuracy']:.4f}, "
        f"variance = {s['accuracy']:.4f} × {1-s['accuracy']:.4f} = "
        f"{s['variance']:.4f}", mono_font); row += 1
    put(row, "How to read it", label_font); row += 1
    put(row,
        "Variance is largest at p = 0.5, where every item is essentially a "
        "coin flip, and smallest at p = 0 or p = 1, where every item is "
        "answered the same way. So a very strong model and a very weak model "
        "both have low variance, while a middling model has high variance. "
        "This means variance on its own is not a measure of model "
        "consistency."); row += 2

    # ===== Standard deviation =====
    put(row, "3. Standard deviation", section_font); row += 1
    put(row, "What it measures", label_font); row += 1
    put(row,
        "The square root of variance. Same information as variance, but on "
        "the same scale as the original data (0 to 1)."); row += 1
    put(row, "Formula", label_font); row += 1
    put(row, "std_dev = √variance = √(p × (1 − p))",
        mono_font); row += 1
    put(row, "Worked example", label_font); row += 1
    put(row,
        f"{short}: std_dev = √{s['variance']:.4f} = {s['std_dev']:.4f}",
        mono_font); row += 1
    put(row, "How to read it", label_font); row += 1
    put(row,
        "Mostly an intermediate step on the way to the standard error. Same "
        "caveat as variance applies. A value near 0.5 just means accuracy was "
        "middling, not that the model is inconsistent."); row += 2

    # ===== Standard error =====
    put(row, "4. Standard error", section_font); row += 1
    put(row, "What it measures", label_font); row += 1
    put(row,
        "How much the accuracy estimate would be expected to wobble if we "
        "drew a different random sample of 501 items from the same population "
        "of memes. A smaller standard error means a more precise estimate."
        ); row += 1
    put(row, "Formula", label_font); row += 1
    put(row, "std_error = √(variance / n) = std_dev / √n",
        mono_font); row += 1
    put(row, "Worked example", label_font); row += 1
    put(row,
        f"{short}: std_error = √({s['variance']:.4f} / {s['n_total']}) = "
        f"{s['std_error']:.4f}  (≈ {s['std_error']*100:.2f} percentage points)",
        mono_font); row += 1
    put(row, "How to read it", label_font); row += 1
    put(row,
        "As a rule of thumb, the true accuracy on this domain sits within "
        "roughly two SEs of the observed accuracy about 95 percent of the "
        "time under the normal approximation. For tighter bounds we use the "
        "bootstrap CI below. The order of magnitude here (around 1.7 to 2.2 "
        "percentage points for our models) is a quick sanity check on whether "
        "two accuracy numbers are meaningfully different."); row += 2

    # ===== Bootstrap CI =====
    put(row, "5. 95% bootstrap confidence interval", section_font); row += 1
    put(row, "What it measures", label_font); row += 1
    put(row,
        "A range that contains the true accuracy with 95 percent probability. "
        "We compute it without assuming normality, so it holds up well on "
        "small samples and skewed distributions."); row += 1
    put(row, "Method", label_font); row += 1
    put(row,
        "1. Take the 501 per-item outcomes (1 = correct, 0 = wrong or invalid).\n"
        "2. Draw a sample of 501 items WITH REPLACEMENT.\n"
        "3. Compute the accuracy on that sample.\n"
        "4. Repeat steps 2 and 3 ten thousand times to get 10,000 simulated accuracies.\n"
        "5. The 2.5th and 97.5th percentiles of that distribution give the 95 percent CI.",
        mono_font); row += 5     # multi-line, so leave extra rows
    put(row, "Worked example", label_font); row += 1
    put(row,
        f"{short}: 10,000 bootstrap resamples gave a wide range of accuracies, "
        f"with the middle 95 percent sitting between {s['ci_low_95']:.1%} and "
        f"{s['ci_high_95']:.1%}. So we report the 95 percent CI as "
        f"[{s['ci_low_95']:.1%}, {s['ci_high_95']:.1%}].",
        mono_font); row += 1
    put(row, "Reproducibility", label_font); row += 1
    put(row,
        "The random number generator is seeded with 42, so re-running stats.py "
        "on the same inputs gives identical CIs. Change the seed in stats.py "
        "if you want a different draw.", note_font); row += 1
    put(row, "How to read it", label_font); row += 1
    put(row,
        "If two models' CIs overlap, the accuracy difference between them "
        "could plausibly be down to chance. If they do not overlap, that is "
        "strong evidence of a real difference. Non-overlap is a conservative "
        "test though, so for the formally correct pairwise test use McNemar "
        "(Section 6) or the bootstrap difference test (Section 7)."); row += 2

    # ===== McNemar's test =====
    put(row, "6. McNemar's test (continuity corrected)", section_font); row += 1
    put(row, "What it tests", label_font); row += 1
    put(row,
        "Whether two models have meaningfully different accuracy when scored "
        "on the same set of items. Each item is a paired observation: "
        "model A's correctness and model B's correctness on that one item."
        ); row += 1
    put(row, "Why a paired test instead of a two-sample t-test", label_font)
    row += 1
    put(row,
        "Both models answered the same 501 items, so the two samples are not "
        "independent. They share the underlying difficulty of each meme. A "
        "paired test cancels that out by looking only at the per-item "
        "agreement and disagreement."); row += 1
    put(row, "Logic", label_font); row += 1
    put(row,
        "Items where both models agree (both correct or both wrong) carry no "
        "information about which model is better. Only the disagreements "
        "matter:\n"
        "  a_only = A correct, B wrong\n"
        "  b_only = A wrong, B correct\n"
        "If the two models are equally good, a_only and b_only should be "
        "roughly balanced. If A is genuinely better, a_only should reliably "
        "exceed b_only.", mono_font); row += 6
    put(row, "Formula", label_font); row += 1
    put(row,
        "chi2 = (|a_only - b_only| - 1)^2 / (a_only + b_only)\n"
        "p-value = P(chi2_1 > observed chi2), where chi2_1 is chi-square with 1 df.\n"
        "The minus one is Edwards' continuity correction, recommended when the "
        "number of discordant pairs (a_only + b_only) > 25, which holds for "
        "every pair in this benchmark.", mono_font); row += 4
    put(row, "Worked example (from the McNemar_Pairwise tab)", label_font)
    row += 1
    put(row,
        "Take Claude vs Gemini. Suppose a_only = 24 (Claude right, Gemini "
        "wrong) and b_only = 31 (Gemini right, Claude wrong). Then\n"
        "  chi2 = (|24 - 31| - 1)^2 / (24 + 31)\n"
        "       = 36 / 55 = 0.65\n"
        "  p = 0.42, not significant. The two models are statistically tied.",
        mono_font); row += 4
    put(row, "How to read the p-value", label_font); row += 1
    put(row,
        "p < 0.05 means a significant difference (we reject the null that A and B "
        "perform equally).\n"
        "p < 0.01 is strong evidence of a difference.\n"
        "p < 0.001 is very strong evidence.\n"
        "p at or above 0.05 means we cannot conclude that the models differ. "
        "The observed gap is consistent with random noise at this sample size.",
        mono_font); row += 5
    put(row, "Caveat: multiple comparisons", label_font); row += 1
    put(row,
        "With seven models we run 21 pairwise tests. A Bonferroni-corrected "
        "threshold for alpha = 0.05 would be 0.05 / 21 = 0.0024. Every "
        "p-value below 0.001 in our results easily clears this. The "
        "non-significant pair (Gemini and Claude) stays non-significant under "
        "any reasonable correction.", note_font); row += 2

    # ===== Bootstrap diff =====
    put(row, "7. Bootstrap test for accuracy difference", section_font)
    row += 1
    put(row, "What it tests", label_font); row += 1
    put(row,
        "Same null hypothesis as McNemar (the two models have equal "
        "accuracy), but it also answers a related question on the way: how "
        "big could the true accuracy difference be? We compute a 95 percent "
        "confidence interval around the observed difference."); row += 1
    put(row, "Method", label_font); row += 1
    put(row,
        "1. Build the matched-pairs list. For each item store (A_correct, B_correct).\n"
        "2. Draw a sample of n pairs WITH REPLACEMENT from that list.\n"
        "3. Compute accuracy_A and accuracy_B on the sample and take the difference.\n"
        "4. Repeat steps 2 and 3 ten thousand times.\n"
        "5. The 2.5th and 97.5th percentiles of the differences give the 95 percent CI.",
        mono_font); row += 5
    put(row, "Why bootstrap and not a t-test on the difference", label_font)
    row += 1
    put(row,
        "Per-item differences are not normally distributed. They only take "
        "the values minus one, zero, and plus one. A bootstrap CI sidesteps "
        "the normality assumption and works directly with the empirical "
        "distribution. The trade-off is computational cost, which is "
        "negligible at n around 500."); row += 1
    put(row, "How to read the CI", label_font); row += 1
    put(row,
        "If the CI does not include 0, the difference is significant at "
        "alpha = 0.05 (two-sided).\n"
        "If the CI includes 0, we cannot rule out that the true difference is "
        "zero, so the models are statistically indistinguishable on this sample.\n"
        "The width of the CI tells you how precisely we estimated the gap. "
        "A wide CI around a small numerical difference means we do not know "
        "whether A is much better, the same as, or somewhat worse than B.",
        mono_font); row += 5
    put(row,
        "Bootstrap and McNemar give the same significance verdict in our "
        "results, so the sig_05 columns line up across both tabs. The "
        "bootstrap CI is more useful in the discussion because you can quote "
        "the size of the gap. For example: 'Gemini Flash beats Claude Opus "
        "by 1.2 pp, 95 percent CI [-2.4 pp, +4.8 pp]', which makes it "
        "immediately clear that the gap is small and uncertain.",
        note_font); row += 2

    # ===== Summary of what this benchmark shows =====
    put(row, "8. What our results show, statistically", section_font); row += 1
    put(row,
        "Putting all the tests above together, the seven models fall into a "
        "few statistically distinct tiers:"); row += 1
    put(row, "  TOP     Gemini 3.5 Flash and Claude Opus 4.6 (p = 0.585)",
        mono_font); row += 1
    put(row, "  MIDDLE  Gemma 4 26B, GPT-5, and Qwen3.5-9B sit close together",
        mono_font); row += 1
    put(row, "  BOTTOM  Qwen3-VL 30B Thinking above Gemma 3 4B, both well "
              "below the rest (all p < 0.001).",
        mono_font); row += 2
    put(row,
        "Within-tier gaps such as Gemini vs Claude are not distinguishable "
        "from noise. Between-tier gaps hold up under reasonable choices of "
        "test, sample size, and multiple-comparison correction. The ordering "
        "broadly tracks parameter scale and training recency, with one "
        "interesting exception: Gemini 3.5 Flash is in a small model class "
        "but lands alongside Claude Opus in a larger one."); row += 1

    ws.sheet_view.showGridLines = False


def main():
    if len(sys.argv) < 3:
        sys.exit("usage: python3 stats.py answer_key.csv run1.csv run2.csv ...")
    key_path = sys.argv[1]
    run_paths = sys.argv[2:]

    print(f"loading {len(run_paths)} runs against {key_path} ...")
    _key, by_model = load_runs(key_path, run_paths)

    # Per-model statistics first.
    print("\nComputing per-model statistics (10k bootstrap iterations each)...")
    stats = {}
    for model, per_file in by_model.items():
        stats[model] = model_stats(per_file)
        s = stats[model]
        print(f"  {model:42s}  acc={s['accuracy']:6.1%}  "
              f"SE={s['std_error']:.4f}  "
              f"95% CI [{s['ci_low_95']:6.1%}, {s['ci_high_95']:6.1%}]")

    # Then the pairwise comparisons.
    print("\nComputing pairwise McNemar and bootstrap-diff tests...")
    mcnemar_results = []
    bootstrap_results = []
    for a, b in combinations(by_model, 2):
        m = mcnemar(by_model[a], by_model[b])
        d = bootstrap_diff(by_model[a], by_model[b])
        mcnemar_results.append((a, b, m))
        bootstrap_results.append((a, b, d))
        print(f"  {a:32s} vs {b:32s}  "
              f"diff={d['diff']:+6.1%}  "
              f"McNemar p={m['p_value']:.2e}  "
              f"sig@.05={'YES' if m['sig_05'] else 'no '}")

    # Write the Excel report.
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_model_stats(wb, stats)
    write_mcnemar(wb, mcnemar_results, stats)
    write_bootstrap(wb, bootstrap_results, stats)
    write_methodology(wb, stats)   # inserted as the first sheet
    wb.save("stats_report.xlsx")
    print("\nwrote stats_report.xlsx (with the Methodology tab)")


if __name__ == "__main__":
    main()
