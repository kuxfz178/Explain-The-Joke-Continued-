#!/usr/bin/env python3
"""Builds the MCQ benchmark files from our annotation spreadsheet.

Reads the Annotations sheet, cleans up the encoding artefacts we kept seeing
in the Reddit exports, and shuffles the four options so the correct answer
lands evenly across A, B, C, and D across the 501 items.

Three files come out:

  annotations_clean.csv  cleaned source data, one row per item
  model_input.csv        what the models actually see (File + option_a..d)
  answer_key.csv         correct letter, literal letter, and the labels

The shuffle is deterministic. Each item is seeded from its own File id plus
the global SEED below, so re-running the script on an edited spreadsheet
keeps every untouched item in its original order.
"""
import csv
import difflib
import re

import ftfy
import openpyxl

SRC = "source.xlsx"
SEED = 20260522            # only change this if you want a new shuffle
NEAR_MATCH = 0.90          # similarity at which we treat Literal as the same as an explanation

COL = dict(File=1, Assigned_to=2, Status=3, Correct=4, Exp1=5, Exp2=6,
           Exp3=7, Literal=8, Humor=9, HumorSec=10, Context=11, ContextSec=12)
XESC = re.compile(r"_x([0-9A-Fa-f]{4})_")


def clean(value):
    """Fix Excel control-char escapes and mojibake, then trim whitespace.

    We use ftfy.fix_encoding rather than fix_text so that actual mojibake gets
    repaired but real content characters (curly quotes, em-dashes, non-breaking
    and zero-width spaces) are kept as is for the audit step to flag.
    """
    s = "" if value is None else str(value)
    s = XESC.sub(lambda m: chr(int(m.group(1), 16)), s)   # turn _x009d_ back into the actual character
    s = ftfy.fix_encoding(s)                              # repair mojibake
    return s.strip()


def loose(s):
    """Lowercase, whitespace-collapsed key for matching Literal to an explanation."""
    s = " ".join((s or "").split()).lower()
    for q in "‘’":
        s = s.replace(q, "'")
    for q in "“”":
        s = s.replace(q, '"')
    return s.replace("—", "-").replace("–", "-")


# Straight and curly double quotes we treat as wrappers.
WRAP_QUOTES = '"“”'


def strip_wrapping_quotes(s):
    """Drop one wrapping pair of double quotes if the text both starts and ends
    with one. Reddit comments often come pasted in this way. If only one side
    has a quote we leave it alone."""
    s = s.strip()
    if len(s) >= 2 and s[0] in WRAP_QUOTES and s[-1] in WRAP_QUOTES:
        s = s[1:-1].strip()
    return s


def stable_rng(*parts):
    """Deterministic RNG seeded from SEED plus the given parts."""
    import random
    return random.Random("|".join([str(SEED)] + [str(p) for p in parts]))


def load_items():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Annotations"]
    items = []
    for r in range(2, ws.max_row + 1):
        fid = ws.cell(row=r, column=COL["File"]).value
        if fid in (None, ""):
            continue
        raw = {k: ws.cell(row=r, column=c).value for k, c in COL.items()}
        item = {"sheet_row": r, "File": str(fid).strip()}
        for k in ("Correct", "Exp1", "Exp2", "Exp3", "Literal"):
            item[k] = clean(raw[k])
            item[k + "_raw"] = "" if raw[k] is None else str(raw[k])
        item["Humor"] = clean(raw["Humor"])
        item["HumorSec"] = clean(raw["HumorSec"])
        item["Context"] = clean(raw["Context"])
        item["ContextSec"] = clean(raw["ContextSec"])
        # Flag encoding repairs before stripping the quotes so we can tell the
        # two cleaning steps apart.
        item["_enc_repaired"] = any(
            item[k] != item[k + "_raw"].strip()
            for k in ("Correct", "Exp1", "Exp2", "Exp3", "Literal"))
        # Strip a wrapping quote pair from every text field, not just Correct.
        # Row 327 was the case that broke duplicate detection for us: Correct
        # and Exp1 held the same text, but Exp1 came in with an extra CSV
        # quote pair and looked different to the duplicate check.
        any_stripped = False
        for k in ("Correct", "Exp1", "Exp2", "Exp3", "Literal"):
            stripped = strip_wrapping_quotes(item[k])
            if stripped != item[k]:
                any_stripped = True
            item[k] = stripped
        item["_quotes_stripped"] = any_stripped
        items.append(item)
    return items


def identify_literal(item):
    """Figure out which of the three explanations is the literal description.

    Returns (exp_key, flag). The flag tells us if anything looked off so the
    audit can pick it up later.
    """
    lit = item["Literal"]
    exps = {k: item[k] for k in ("Exp1", "Exp2", "Exp3")}
    if not loose(lit):
        return None, "missing_literal"
    if loose(lit) == loose(item["Correct"]):
        return None, "literal_equals_correct"
    exact = [k for k, v in exps.items() if loose(v) and loose(v) == loose(lit)]
    if len(exact) == 1:
        return exact[0], ""
    if len(exact) > 1:
        return exact[0], "literal_matches_multiple"
    # no exact match -> nearest by similarity
    best_k, best_r = None, 0.0
    for k, v in exps.items():
        r = difflib.SequenceMatcher(None, loose(lit), loose(v)).ratio()
        if r > best_r:
            best_k, best_r = k, r
    if best_r >= NEAR_MATCH:
        return best_k, "literal_paraphrased"
    return None, "literal_no_match"


def main():
    items = load_items()
    n = len(items)
    print(f"Loaded {n} items from {SRC}")

    # ---- balanced, fixed assignment of the correct-answer letter ----------
    letters = ["A", "B", "C", "D"]
    base, rem = divmod(n, 4)
    pool = []
    for i, L in enumerate(letters):
        pool += [L] * (base + (1 if i < rem else 0))
    stable_rng("correct-letter-pool").shuffle(pool)
    by_file = sorted(items, key=lambda it: it["File"])
    correct_letter = {it["File"]: pool[i] for i, it in enumerate(by_file)}

    clean_rows, model_rows, key_rows = [], [], []
    flag_counts = {}

    for it in items:
        fid = it["File"]
        flags = []

        # encoding repair flags
        if it["_enc_repaired"]:
            flags.append("encoding_repaired")
        if it["_quotes_stripped"]:
            flags.append("wrapping_quotes_stripped")
        for k in ("Correct", "Exp1", "Exp2", "Exp3", "Literal"):
            if "�" in it[k]:
                flags.append("encoding_unrecoverable")  # data lost - retype
                break

        # which explanation is the literal one
        lit_key, lit_flag = identify_literal(it)
        if lit_flag:
            flags.append(lit_flag)

        # duplicate options within the item
        opt_text = {"Correct": it["Correct"], "Exp1": it["Exp1"],
                    "Exp2": it["Exp2"], "Exp3": it["Exp3"]}
        seen = {}
        for k, v in opt_text.items():
            if loose(v):
                seen.setdefault(loose(v), []).append(k)
        if any(len(v) > 1 for v in seen.values()):
            flags.append("duplicate_options")
        if any(not loose(v) for v in opt_text.values()):
            flags.append("blank_option")

        # ---- place the four options into letter slots ----------------------
        cl = correct_letter[fid]
        slots = {cl: ("Correct", it["Correct"])}
        free = [L for L in letters if L != cl]
        distractors = [("Exp1", it["Exp1"]), ("Exp2", it["Exp2"]),
                       ("Exp3", it["Exp3"])]
        stable_rng("distractor-order", fid).shuffle(distractors)
        for L, (src, txt) in zip(free, distractors):
            slots[L] = (src, txt)

        literal_letter = ""
        if lit_key:
            literal_letter = next(L for L, (src, _) in slots.items()
                                  if src == lit_key)

        for f in flags:
            flag_counts[f] = flag_counts.get(f, 0) + 1

        clean_rows.append({
            "File": fid, "Correct answer": it["Correct"],
            "Explanation 1": it["Exp1"], "Explanation 2": it["Exp2"],
            "Explanation 3": it["Exp3"], "Literal_Description": it["Literal"],
            "Humor Style": it["Humor"],
            "Humor Style Secondary": it["HumorSec"],
            "Contextual Knowledge": it["Context"],
            "Contextual Knowledge Secondary": it["ContextSec"],
        })
        model_rows.append({
            "File": fid,
            "option_a": slots["A"][1], "option_b": slots["B"][1],
            "option_c": slots["C"][1], "option_d": slots["D"][1],
        })
        key_rows.append({
            "File": fid,
            "correct_letter": cl,
            "literal_letter": literal_letter or "unknown",
            "literal_source_column": lit_key or "",
            "humor_style": it["Humor"],
            "humor_style_secondary": it["HumorSec"],
            "contextual_knowledge": it["Context"],
            "contextual_knowledge_secondary": it["ContextSec"],
            "correct_answer_char_len": len(it["Correct"]),
            "data_quality_flags": ";".join(sorted(set(flags))),
            "sheet_row": it["sheet_row"],
        })

    # ---- write the three files -------------------------------------------
    def dump(path, rows):
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()),
                               quoting=csv.QUOTE_ALL)
            w.writeheader()
            w.writerows(rows)
        print(f"  wrote {path}  ({len(rows)} rows)")

    print("\nOutput files:")
    dump("annotations_clean.csv", clean_rows)
    dump("model_input.csv", model_rows)
    dump("answer_key.csv", key_rows)

    # ---- summary ----------------------------------------------------------
    from collections import Counter
    dist = Counter(r["correct_letter"] for r in key_rows)
    litknown = sum(1 for r in key_rows if r["literal_letter"] != "unknown")
    print(f"\nCorrect-answer letter distribution: "
          f"{dict(sorted(dist.items()))}  (target ~{n/4:.0f} each)")
    print(f"Items with an identifiable literal option: {litknown}/{n}  "
          f"({n - litknown} cannot be literal-scored)")
    print("\nData-quality flags raised:")
    for f in sorted(flag_counts, key=lambda k: -flag_counts[k]):
        print(f"  {flag_counts[f]:4d}  {f}")


if __name__ == "__main__":
    main()
